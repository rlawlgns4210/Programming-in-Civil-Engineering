import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill

def save_output_info(main_bridge, side_bridge=None, addition_bridge=None, num_bridges=0, num_side_bridge=0, num_addition_bridge=0, remaining_bridge_length=0, steel_material=None, num_panels=0):
    # 주교 정보
    main_bridge_data = [
        ["길이(m)", 60, num_panels * 10 + remaining_bridge_length if side_bridge else 0, 0 if not addition_bridge else (addition_bridge.num_panels * 10)],
        ["개수", num_bridges, num_side_bridge, num_addition_bridge],
        ["자중(kN/m)", main_bridge.calculation_details['w1'], side_bridge.calculation_details['w1'] if side_bridge else 0, addition_bridge.calculation_details['w1'] if addition_bridge else 0],
        ["분포하중(kN/m)", main_bridge.calculation_details['w2'], side_bridge.calculation_details['w2'] if side_bridge else 0, addition_bridge.calculation_details['w2'] if addition_bridge else 0],
        ["총 하중(kN/m)", main_bridge.calculation_details['W'], side_bridge.calculation_details['W'] if side_bridge else 0, addition_bridge.calculation_details['W'] if addition_bridge else 0],
        ["각도 a(˚)", main_bridge.calculation_details['a'], side_bridge.calculation_details['a'] if side_bridge else 0, addition_bridge.calculation_details['a'] if addition_bridge else 0],
        ["각도 b(˚)", "", side_bridge.calculation_details['b'] if side_bridge else "", ""]
    ]
    # 데이터프레임 생성
    df1 = pd.DataFrame(main_bridge_data, columns=["항목", "주교", "측교", "추가교량"])
    
    # 항복 강도 값 설정
    yield_strengths = {
        "SS235": 235,
        "SS275": 275,
        "SS315": 315,
        "SS410": 410,
        "SS450": 450,
        "SS550": 550
    }
    yield_strength = yield_strengths.get(steel_material, None)
    
    if yield_strength is None:
        raise ValueError(f"Unknown steel material: {steel_material}")
    
    # MainBridge 시트 데이터 생성
    main_bridge_forces = main_bridge.member_forces
    main_bridge_stresses = main_bridge.stress_forces
    
    mainbridge_data = []
    for i in range(len(main_bridge_forces)):
        force = main_bridge_forces[i]
        stress = main_bridge_stresses[i] # N/mm^2
        if abs(stress) < yield_strength:
            below_yield_strength = "O"
        elif abs(stress) >= yield_strength:
            below_yield_strength = "X"
        # below_yield_strength = "O" if stress < yield_strength else "X"
        mainbridge_data.append([f"F{i+1}", force, stress, yield_strength, below_yield_strength])

    # 데이터프레임 생성
    df2 = pd.DataFrame(mainbridge_data, columns=["부재력번호", "부재력", "부재응력", "항복강도", "부재응력<항복강도"])
    
    # SideBridge 시트 데이터 생성 (있는 경우)
    if side_bridge:
        side_bridge_forces = side_bridge.member_forces
        side_bridge_stresses = side_bridge.stress_forces
        sidebridge_data = []
        num_forces = len(side_bridge_forces)
        for i in range(num_forces):
            if i < 2:
                force_label = f"Fs{i+1}"
            else:
                force_label = f"F{i-1}"
            force = side_bridge_forces[i]
            stress = side_bridge_stresses[i]  # N/mm^2
            if abs(stress) < yield_strength:
                below_yield_strength = "O"
            elif abs(stress) >= yield_strength:
                below_yield_strength = "X"
            # below_yield_strength = "O" if stress < yield_strength else "X"
            sidebridge_data.append([force_label, force, stress, yield_strength, below_yield_strength])
        df3 = pd.DataFrame(sidebridge_data, columns=["부재력번호", "부재력", "부재응력", "항복강도", "부재응력<항복강도"])
    
    # AdditionBridge 시트 데이터 생성 (있는 경우)
    if addition_bridge:
        addition_bridge_forces = addition_bridge.member_forces
        addition_bridge_stresses = addition_bridge.stress_forces
        additionbridge_data = []
        num_forces = len(addition_bridge_forces)
        for i in range(num_forces):
            if i < 2:
                force_label = f"F{i+1}"
            else:
                force_label = f"F{i-1}"
            force = addition_bridge_forces[i]
            stress = addition_bridge_stresses[i]  # N/mm^2
            if abs(stress) < yield_strength:
                below_yield_strength = "O"
            elif abs(stress) >= yield_strength:
                below_yield_strength = "X"
            # below_yield_strength = "O" if abs(stress) < yield_strength else "X"
            additionbridge_data.append([force_label, force, stress, yield_strength, below_yield_strength])
        df4 = pd.DataFrame(additionbridge_data, columns=["부재력번호", "부재력", "부재응력", "항복강도", "부재응력<항복강도"])
    
    # 엑셀 파일로 저장
    with pd.ExcelWriter('OutputInform.xlsx', engine='openpyxl') as writer:
        # 첫 번째 시트 저장
        df1.to_excel(writer, sheet_name='bridgeinform', index=False, header=True)
        
        # 두 번째 시트 저장
        df2.to_excel(writer, sheet_name='mainbridge', index=False, header=True)
        
        # 세 번째 시트 저장 (있는 경우)
        if side_bridge:
            df3.to_excel(writer, sheet_name='sidebridge', index=False, header=True)
        
        # 네 번째 시트 저장 (있는 경우)
        if addition_bridge:
            df4.to_excel(writer, sheet_name='additionbridge', index=False, header=True)

        # 스타일링 추가
        workbook = writer.book
        bridgeinform_sheet = workbook['bridgeinform']
        mainbridge_sheet = workbook['mainbridge']
        if side_bridge:
            sidebridge_sheet = workbook['sidebridge']
        if addition_bridge:
            additionbridge_sheet = workbook['additionbridge']

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="006600", end_color="006600", fill_type="solid")

        for col in bridgeinform_sheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=4):
            for cell in col:
                cell.font = header_font
                cell.fill = header_fill

        for col in mainbridge_sheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=5):
            for cell in col:
                cell.font = header_font
                cell.fill = header_fill

        if side_bridge:
            for col in sidebridge_sheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=5):
                for cell in col:
                    cell.font = header_font
                    cell.fill = header_fill

        if addition_bridge:
            for col in additionbridge_sheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=5):
                for cell in col:
                    cell.font = header_font
                    cell.fill = header_fill

        # Save the styled workbook
        writer.book.save('OutputInform.xlsx')
